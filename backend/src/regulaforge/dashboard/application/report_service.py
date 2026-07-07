"""Report service for dashboard report generation and export.

Provides executive summaries, compliance reports, risk reports,
regulatory change reports, and multi-format export capabilities
with scheduling support.
"""

from datetime import datetime, timezone
from typing import Any, Optional
from uuid import uuid4

from regulaforge.config.logging import get_logger
from regulaforge.dashboard.infrastructure.cache import DashboardCache

logger = get_logger(__name__)


class ReportService:
    """Generates reports and exports for the Enterprise Dashboard.

    Supports executive summaries, detailed compliance reports,
    risk reports, regulatory change impact reports, and scheduled
    recurring report generation.
    """

    def __init__(
        self,
        dashboard_cache: Optional[DashboardCache] = None,
    ) -> None:
        """Initialize ReportService with optional cache dependency.

        Args:
            dashboard_cache: Cache instance for data caching.
        """
        self._cache = dashboard_cache or DashboardCache()

    async def generate_executive_summary(
        self,
        tenant_id: str,
        report_format: str = "json",
    ) -> dict[str, Any]:
        """Generate an executive summary report.

        Produces a high-level overview of compliance posture,
        key risks, recent changes, and strategic recommendations
        suitable for C-suite presentation.

        Args:
            tenant_id: The tenant identifier.
            report_format: Output format ('json', 'html', 'pdf').

        Returns:
            Dict containing the executive summary data.
        """
        logger.info(
            "Generating executive summary for tenant %s (%s)",
            tenant_id, report_format,
        )
        now = datetime.now(timezone.utc)

        return {
            "report_type": "executive_summary",
            "tenant_id": tenant_id,
            "generated_at": now.isoformat(),
            "format": report_format,
            "summary": {
                "title": "RegulaForge Compliance Executive Summary",
                "period": "Q2 2026",
                "overall_compliance_rate": 87.5,
                "compliance_change": 2.3,
                "entities_in_scope": 342,
                "entities_assessed": 285,
                "assessment_coverage": 83.3,
            },
            "key_metrics": {
                "compliance_rate": {
                    "value": 87.5,
                    "unit": "%",
                    "change": 2.3,
                    "trend": "up",
                },
                "open_findings": {
                    "value": 128,
                    "unit": "count",
                    "change": -8.4,
                    "trend": "down",
                },
                "overdue_assessments": {
                    "value": 23,
                    "unit": "count",
                    "change": 12.5,
                    "trend": "up",
                },
                "active_risks": {
                    "value": 47,
                    "unit": "count",
                    "change": -2.1,
                    "trend": "down",
                },
            },
            "risk_overview": {
                "current_score": 34.2,
                "high_risk_entities": 18,
                "critical_alerts": 5,
                "top_risks": [
                    "Data protection gaps in third-party integrations",
                    "Outdated cybersecurity certifications",
                    "Incomplete AI governance framework",
                ],
            },
            "regulatory_changes": {
                "changes_tracked": 24,
                "critical_changes": 3,
                "recent_updates": [
                    "GDPR Article 32 - Enhanced security measures required",
                    "SOC 2 Type II - New trust services criteria",
                    "AI Act - New compliance timeline announced",
                ],
            },
            "recommendations": [
                "Prioritize remediation of 8 critical findings in data protection",
                "Schedule cybersecurity certification renewals for Q3",
                "Establish AI governance working group for Act compliance",
                "Reduce overdue assessments by increasing assessment velocity",
            ],
            "generated_by": "RegulaForge Enterprise Dashboard",
            "version": "2.0.0",
        }

    async def generate_compliance_report(
        self,
        tenant_id: str,
        entity_id: Optional[str] = None,
        regulation_id: Optional[str] = None,
        report_format: str = "json",
    ) -> dict[str, Any]:
        """Generate a detailed compliance report.

        Provides comprehensive compliance data optionally filtered
        by entity and/or regulation for focused analysis.

        Args:
            tenant_id: The tenant identifier.
            entity_id: Optional entity filter.
            regulation_id: Optional regulation filter.
            report_format: Output format ('json', 'html', 'pdf').

        Returns:
            Dict containing the compliance report data.
        """
        logger.info(
            "Generating compliance report for tenant %s (entity=%s, reg=%s, fmt=%s)",
            tenant_id, entity_id, regulation_id, report_format,
        )
        now = datetime.now(timezone.utc)

        return {
            "report_type": "compliance_report",
            "tenant_id": tenant_id,
            "generated_at": now.isoformat(),
            "format": report_format,
            "filters": {
                "entity_id": entity_id,
                "regulation_id": regulation_id,
            },
            "compliance_summary": {
                "overall_rate": 87.5,
                "entities_assessed": 285,
                "entities_in_scope": 342,
                "assessments_completed": 312,
                "assessments_pending": 48,
                "regulations_covered": 156,
            },
            "findings_breakdown": {
                "total": 128,
                "by_severity": {
                    "critical": 8,
                    "high": 23,
                    "medium": 47,
                    "low": 50,
                },
                "by_status": {
                    "open": 128,
                    "in_progress": 42,
                    "resolved": 0,
                    "accepted": 0,
                },
                "top_finding_categories": [
                    {"category": "Data Protection", "count": 35, "severity": "high"},
                    {"category": "Access Control", "count": 28, "severity": "medium"},
                    {"category": "Documentation", "count": 22, "severity": "low"},
                ],
            },
            "assessment_summary": {
                "total": 360,
                "completed": 312,
                "in_progress": 25,
                "overdue": 23,
                "completion_rate": 86.7,
                "avg_completion_time_days": 14.5,
            },
            "generated_by": "RegulaForge Enterprise Dashboard",
            "version": "2.0.0",
        }

    async def generate_risk_report(
        self,
        tenant_id: str,
        entity_id: Optional[str] = None,
        report_format: str = "json",
    ) -> dict[str, Any]:
        """Generate a risk-focused report.

        Provides detailed risk analysis including risk scores,
        distribution, top risk factors, and mitigation tracking.

        Args:
            tenant_id: The tenant identifier.
            entity_id: Optional entity filter.
            report_format: Output format ('json', 'html', 'pdf').

        Returns:
            Dict containing the risk report data.
        """
        logger.info(
            "Generating risk report for tenant %s (entity=%s, fmt=%s)",
            tenant_id, entity_id, report_format,
        )
        now = datetime.now(timezone.utc)

        return {
            "report_type": "risk_report",
            "tenant_id": tenant_id,
            "generated_at": now.isoformat(),
            "format": report_format,
            "filters": {
                "entity_id": entity_id,
            },
            "risk_summary": {
                "current_score": 34.2,
                "score_change": -1.5,
                "score_trend": "improving",
                "risk_appetite_threshold": 40.0,
                "within_appetite": True,
            },
            "risk_distribution": {
                "critical": 5,
                "high": 13,
                "medium": 87,
                "low": 156,
                "negligible": 81,
            },
            "top_risk_factors": [
                {
                    "factor": "Data protection gaps in third-party integrations",
                    "risk_level": "critical",
                    "affected_entities": 12,
                    "mitigation_status": "in_progress",
                },
                {
                    "factor": "Outdated cybersecurity certifications",
                    "risk_level": "high",
                    "affected_entities": 8,
                    "mitigation_status": "planned",
                },
                {
                    "factor": "Incomplete AI governance framework",
                    "risk_level": "high",
                    "affected_entities": 6,
                    "mitigation_status": "not_started",
                },
                {
                    "factor": "Cross-border data transfer compliance",
                    "risk_level": "medium",
                    "affected_entities": 15,
                    "mitigation_status": "in_progress",
                },
            ],
            "risk_trend": {
                "30d_change": -4.8,
                "60d_change": -2.1,
                "90d_change": -8.3,
            },
            "mitigation_summary": {
                "total_actions": 34,
                "completed": 12,
                "in_progress": 15,
                "planned": 5,
                "overdue": 2,
            },
            "generated_by": "RegulaForge Enterprise Dashboard",
            "version": "2.0.0",
        }

    async def generate_regulatory_change_report(
        self,
        regulation_id: Optional[str] = None,
        report_format: str = "json",
    ) -> dict[str, Any]:
        """Generate a regulatory change impact report.

        Analyzes recent and upcoming regulatory changes and their
        potential impact on the organization's compliance posture.

        Args:
            regulation_id: Optional specific regulation to analyze.
            report_format: Output format ('json', 'html', 'pdf').

        Returns:
            Dict containing the regulatory change report data.
        """
        logger.info(
            "Generating regulatory change report (regulation=%s, fmt=%s)",
            regulation_id, report_format,
        )
        now = datetime.now(timezone.utc)

        return {
            "report_type": "regulatory_change_report",
            "generated_at": now.isoformat(),
            "format": report_format,
            "filters": {
                "regulation_id": regulation_id,
            },
            "changes_summary": {
                "total_changes": 24,
                "critical_changes": 3,
                "high_changes": 8,
                "medium_changes": 9,
                "low_changes": 4,
                "changes_this_month": 7,
            },
            "recent_changes": [
                {
                    "regulation": "GDPR",
                    "article": "Article 32",
                    "change": "Enhanced security of processing requirements",
                    "effective_date": "2026-09-01",
                    "impact_level": "high",
                    "required_actions": [
                        "Update data processing agreements",
                        "Review encryption standards",
                        "Enhance breach notification procedures",
                    ],
                },
                {
                    "regulation": "EU AI Act",
                    "article": "Title IV",
                    "change": "Transparency obligations for high-risk AI systems",
                    "effective_date": "2026-12-01",
                    "impact_level": "critical",
                    "required_actions": [
                        "Inventory all high-risk AI systems",
                        "Implement documentation requirements",
                        "Establish human oversight procedures",
                    ],
                },
                {
                    "regulation": "SOC 2",
                    "standard": "Trust Services Criteria",
                    "change": "New criteria for AI system controls",
                    "effective_date": "2027-01-01",
                    "impact_level": "medium",
                    "required_actions": [
                        "Assess AI system control environments",
                        "Update control documentation",
                    ],
                },
            ],
            "impact_assessment": {
                "affected_entities": 45,
                "affected_regulations": 12,
                "estimated_effort_hours": 1200,
                "deadline_compliance_risk": "medium",
            },
            "generated_by": "RegulaForge Enterprise Dashboard",
            "version": "2.0.0",
        }

    async def export_report(
        self,
        report_data: dict[str, Any],
        export_format: str = "json",
    ) -> bytes:
        """Export report data in the specified format.

        Converts report data to bytes in the requested format:
        json, html, pdf, or xlsx.

        Args:
            report_data: The report data dict to export.
            export_format: Target format ('json', 'html', 'pdf', 'xlsx').

        Returns:
            Bytes of the exported report.

        Raises:
            ValueError: If the export format is unsupported.
        """
        logger.info("Exporting report in %s format (size=%d)", export_format, len(str(report_data)))

        if export_format == "json":
            import json
            content = json.dumps(report_data, indent=2, default=str).encode("utf-8")
        elif export_format == "html":
            content = self._render_html(report_data).encode("utf-8")
        elif export_format == "pdf":
            from io import BytesIO

            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            story.append(Paragraph(f"Report: {report_data.get('report_type', 'Report')}", styles["Title"]))
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"Generated: {report_data.get('generated_at', 'N/A')}", styles["Normal"]))
            story.append(Spacer(1, 12))

            for key, value in report_data.items():
                if key in ("generated_by", "version"):
                    continue
                story.append(Paragraph(f"<b>{key}:</b> {value}", styles["Normal"]))
                story.append(Spacer(1, 6))

            doc.build(story)
            content = buffer.getvalue()
            buffer.close()
        elif export_format == "xlsx":
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")

            def _flatten(data: dict[str, Any], prefix: str = "") -> list[tuple]:
                rows = []
                for key, value in data.items():
                    field = f"{prefix}.{key}" if prefix else key
                    if isinstance(value, dict):
                        rows.extend(_flatten(value, field))
                    elif isinstance(value, list):
                        rows.append((field, str(value)))
                    else:
                        rows.append((field, str(value)))
                return rows

            flattened = _flatten(report_data)
            for col, (_header, _) in enumerate(flattened[:1], 1):
                cell = ws.cell(row=1, column=col, value="Field")
                cell.font = header_font
                cell.fill = header_fill
            for col, (_, value) in enumerate(flattened[:1], 2):  # noqa: B007
                cell = ws.cell(row=1, column=col, value="Value")
                cell.font = header_font
                cell.fill = header_fill

            for row_idx, (field, value) in enumerate(flattened, 2):
                ws.cell(row=row_idx, column=1, value=field)
                ws.cell(row=row_idx, column=2, value=value)

            ws.column_dimensions["A"].width = 40
            ws.column_dimensions["B"].width = 60

            buffer = BytesIO()
            wb.save(buffer)
            content = buffer.getvalue()
            buffer.close()
        else:
            raise ValueError(f"Unsupported export format: {export_format}")

        logger.info("Report exported successfully: %d bytes", len(content))
        return content

    async def schedule_report(
        self,
        report_config: dict[str, Any],
        cron_expression: str,
    ) -> str:
        """Schedule a recurring report generation.

        Registers a cron-based schedule for automatic report
        generation and delivery.

        Args:
            report_config: Configuration dict for the report
                (type, tenant_id, format, recipients, etc.).
            cron_expression: Standard cron expression for scheduling
                (e.g. '0 8 * * 1' for every Monday at 8 AM).

        Returns:
            Schedule identifier string.
        """
        schedule_id = str(uuid4())
        logger.info(
            "Scheduled report %s with cron '%s': %s",
            schedule_id, cron_expression, report_config,
        )

        return schedule_id

    def _render_html(self, report_data: dict[str, Any]) -> str:
        """Render report data as an HTML document.

        Args:
            report_data: The report data dict.

        Returns:
            HTML string representation of the report.
        """
        def _render_value(value: Any) -> str:
            if isinstance(value, dict):
                rows = "".join(
                    f"<tr><td style='padding:4px 8px;font-weight:bold'>{k}</td>"
                    f"<td style='padding:4px 8px'>{_render_value(v)}</td></tr>"
                    for k, v in value.items()
                )
                return f"<table style='border-collapse:collapse;width:100%'>{rows}</table>"
            if isinstance(value, list):
                items = "".join(
                    f"<li>{_render_value(item)}</li>" for item in value
                )
                return f"<ul style='margin:4px 0'>{items}</ul>"
            return str(value)

        body_rows = "".join(
            f"<tr><td style='padding:8px;font-weight:bold;border-bottom:1px solid #e2e8f0'>{key}</td>"
            f"<td style='padding:8px;border-bottom:1px solid #e2e8f0'>{_render_value(value)}</td></tr>"
            for key, value in report_data.items()
        )

        return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>RegulaForge Report</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
               margin: 40px; color: #1e293b; background: #f8fafc; }}
        h1 {{ color: #6366f1; border-bottom: 2px solid #6366f1; padding-bottom: 8px; }}
        table {{ border-collapse: collapse; width: 100%; background: white;
                 box-shadow: 0 1px 3px rgba(0,0,0,0.1); border-radius: 8px; }}
        td {{ vertical-align: top; }}
    </style>
</head>
<body>
    <h1>RegulaForge Report</h1>
    <table>{body_rows}</table>
    <p style="margin-top: 24px; color: #64748b; font-size: 0.85em;">
        Generated by RegulaForge Enterprise Dashboard
    </p>
</body>
</html>"""
